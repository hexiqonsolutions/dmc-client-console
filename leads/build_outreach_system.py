#!/usr/bin/env python3
"""
Rebuild prospect workbook:
- Drop leads with neither usable email nor usable phone
- Add Select + one-click Email / WhatsApp columns
- Emit HTML Outreach Console for multi-select bulk send
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import sys

ROOT = Path(r"C:\Users\FPIN\Desktop\dm os\leads")
sys.path.insert(0, str(ROOT))
from polish_copy import eye_catchy_subject, polish_all_leads  # noqa: E402

SOURCE = ROOT / "build_navi_mumbai_leads.py"
XLSX_OUT = ROOT / "Navi_Mumbai_Outreach_System.xlsx"
HTML_OUT = ROOT / "Outreach_Console.html"
JSON_OUT = ROOT / "prospects_contactable.json"


def load_leads() -> list[dict]:
    src = SOURCE.read_text(encoding="utf-8")
    start = src.index("LEADS = [")
    end = src.index("\nassert len(LEADS)")
    ns: dict = {}
    exec(src[start:end], ns)
    return ns["LEADS"]


def extract_emails(raw: str) -> list[str]:
    if not raw:
        return []
    low = raw.lower()
    if "not found" in low or "gated" in low or "form only" in low:
        return []
    found = re.findall(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", raw)
    # Drop obviously placeholder / truncated
    clean = []
    for e in found:
        el = e.lower()
        if el.endswith("…") or "…" in e:
            continue
        if el.endswith("@g") or el.endswith("@savlagroup") is False and el.count(".") < 1:
            pass
        if re.search(r"@(g|ymail|yahoo|rediffmail|hotmail|gmail|vsnl)\.?$", el):
            continue
        clean.append(e)
    # Keep unique order
    out = []
    for e in found:
        if e not in out and not e.lower().endswith(("…",)):
            # skip truncated directory emails like neelkanthpolymer@g…
            if "…" in e or e.endswith("@"):
                continue
            out.append(e)
    return out


def extract_mobiles(raw: str) -> list[str]:
    """Return WhatsApp-ready mobiles as 91XXXXXXXXXX."""
    if not raw:
        return []
    low = raw.lower()
    if "not found" in low or "gated" in low or "show number" in low:
        return []
    # Collapse spaces/dashes between digit groups so "98204 75880" / "+91-98204-75880" work
    compact = re.sub(r"[^\d+]", "", raw)
    out: list[str] = []
    for m in re.finditer(r"(?:91)?([6-9]\d{9})", compact):
        num = "91" + m.group(1)
        if num not in out:
            out.append(num)
    return out


def has_any_phone(raw: str) -> bool:
    """True if a real phone appears (mobile or landline), for keep/drop filter."""
    if not raw:
        return False
    low = raw.lower()
    if "not found" in low or "gated" in low or "show number" in low:
        return False
    digits = re.sub(r"\D", "", raw)
    # At least one 10-digit sequence somewhere
    return bool(re.search(r"[6-9]\d{9}", digits) or re.search(r"0\d{9,11}", digits) or len(digits) >= 10)


def display_phone(e164: str) -> str:
    if e164.startswith("91") and len(e164) == 12:
        return "+91 " + e164[2:7] + " " + e164[7:]
    return e164


def usable(lead: dict) -> bool:
    return bool(extract_emails(lead.get("email", "")) or has_any_phone(lead.get("phone", "")))


def mailto_url(email: str, subject: str, body: str) -> str:
    return f"mailto:{email}?subject={quote(subject)}&body={quote(body)}"


def wa_url(phone_e164: str, text: str) -> str:
    return f"https://wa.me/{phone_e164}?text={quote(text)}"


def email_subject(lead: dict) -> str:
    if lead.get("subject"):
        return lead["subject"]
    return eye_catchy_subject(lead)


def build_contactable(leads: list[dict]) -> list[dict]:
    leads = polish_all_leads(leads)
    rows = []
    for i, lead in enumerate(leads, 1):
        emails = extract_emails(lead.get("email", ""))
        mobiles = extract_mobiles(lead.get("phone", ""))
        raw_phone = (lead.get("phone") or "").strip()
        phone_ok = has_any_phone(raw_phone)
        if not emails and not phone_ok:
            continue
        primary_email = emails[0] if emails else ""
        primary_phone = mobiles[0] if mobiles else ""
        # Display: prefer mobile formatting, else raw public phone string
        if primary_phone:
            phone_display = display_phone(primary_phone)
            phone_all = " / ".join(display_phone(p) for p in mobiles)
        else:
            phone_display = raw_phone if phone_ok else ""
            phone_all = phone_display
        subject = email_subject(lead)
        body = lead["outreach"]
        rows.append(
            {
                **lead,
                "source_id": i,
                "emails": emails,
                "phones": mobiles,
                "primary_email": primary_email,
                "primary_phone": primary_phone,
                "primary_phone_display": phone_display,
                "email_all": "; ".join(emails),
                "phone_all": phone_all,
                "mailto": mailto_url(primary_email, subject, body) if primary_email else "",
                "whatsapp": wa_url(primary_phone, body) if primary_phone else "",
                "email_subject": subject,
            }
        )
    return rows


def style_header(cell, fill, font):
    cell.fill = fill
    cell.font = font
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")


def build_xlsx(rows: list[dict]) -> None:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    link_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    blue_fill = PatternFill("solid", fgColor="DBEAFE")
    p1_fill = PatternFill("solid", fgColor="FEE2E2")
    p2_fill = PatternFill("solid", fgColor="FEF3C7")
    p3_fill = PatternFill("solid", fgColor="E0E7FF")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    # ---------- START HERE ----------
    ws0 = wb.active
    ws0.title = "START HERE"
    ws0["A1"] = "Navi Mumbai Outreach System"
    ws0["A1"].font = Font(bold=True, size=18, color="0F172A")
    lines = [
        "",
        f"Contactable prospects: {len(rows)} (removed anyone with neither email nor phone)",
        "",
        "HOW TO USE (Excel)",
        "1. Open sheet: Contactable Prospects",
        "2. Column A (Select) — put YES on clients you want to message",
        "3. Column B (Email) — click SEND EMAIL for one client (opens Outlook/Gmail mailto)",
        "4. Column C (WhatsApp) — click SEND WHATSAPP (opens wa.me with personalised message)",
        "",
        "BULK / MULTI-SELECT (recommended)",
        "5. Open Outreach_Console.html in Chrome (same folder) for:",
        "   • Click a row to VIEW + EDIT the message, email, and WhatsApp phone",
        "   • Save changes — Email/WhatsApp links update instantly (mailto + wa.me)",
        "   • Select many clients → Email selected (BCC) or WhatsApp queue",
        "",
        "WHERE TO LINK WHATSAPP / EMAIL",
        "• In Outreach_Console.html right panel: fill Email + Phone → Save → Send buttons",
        "• In Excel: SEND EMAIL / SEND WHATSAPP columns (built from Primary Email / Phone)",
        "• No WhatsApp Business API login required — uses your phone/WhatsApp Web + mail app",
        "",
        "TIPS",
        "• WhatsApp needs a 10-digit Indian mobile. Landline-only → use Email or add mobile in Edit panel.",
        "• Always verify before first send — numbers/emails came from public directories.",
        "• Edits in the console are saved in the browser (localStorage) until you clear site data.",
        "",
        "FILES IN THIS FOLDER",
        f"• {XLSX_OUT.name}",
        f"• {HTML_OUT.name}  ← bulk select + send",
        f"• {JSON_OUT.name}",
    ]
    for i, line in enumerate(lines, 2):
        ws0.cell(i, 1, line)
        if line.startswith("HOW") or line.startswith("BULK") or line.startswith("TIPS") or line.startswith("FILES"):
            ws0.cell(i, 1).font = Font(bold=True, size=12, color="0F172A")
    ws0.column_dimensions["A"].width = 110

    # ---------- Contactable Prospects ----------
    ws = wb.create_sheet("Contactable Prospects", 1)
    headers = [
        "Select",
        "Email (1-click)",
        "WhatsApp (1-click)",
        "ID",
        "Company Name",
        "Owner / Contact",
        "Primary Email",
        "Primary Phone",
        "All Emails",
        "All Phones",
        "Website",
        "What's Wrong",
        "Industry",
        "Priority",
        "Locality",
        "Ideal Offer",
        "Email Subject",
        "Outreach Message (personalised)",
        "Has Email",
        "Has WhatsApp",
    ]
    for c, h in enumerate(headers, 1):
        style_header(ws.cell(1, c, h), header_fill, header_font)

    dv = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
    dv.error = "Choose YES or NO"
    dv.prompt = "Select for bulk"
    ws.add_data_validation(dv)

    for idx, row in enumerate(rows, 1):
        r = idx + 1
        ws.cell(r, 1, "NO")
        dv.add(ws.cell(r, 1))
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(r, 1).fill = PatternFill("solid", fgColor="F8FAFC")

        # Email button
        email_cell = ws.cell(r, 2)
        if row["mailto"]:
            email_cell.value = "SEND EMAIL"
            email_cell.hyperlink = row["mailto"]
            email_cell.font = link_font
            email_cell.fill = blue_fill
        else:
            email_cell.value = "—"
            email_cell.font = Font(color="94A3B8")
        email_cell.alignment = Alignment(horizontal="center", vertical="top")

        # WhatsApp button
        wa_cell = ws.cell(r, 3)
        if row["whatsapp"]:
            wa_cell.value = "SEND WHATSAPP"
            wa_cell.hyperlink = row["whatsapp"]
            wa_cell.font = Font(name="Calibri", size=10, color="166534", underline="single")
            wa_cell.fill = green_fill
        else:
            wa_cell.value = "—"
            wa_cell.font = Font(color="94A3B8")
        wa_cell.alignment = Alignment(horizontal="center", vertical="top")

        values = [
            row["source_id"],
            row["company"],
            row["owner"],
            row["primary_email"] or "—",
            row["primary_phone_display"] or "—",
            row["email_all"] or "—",
            row["phone_all"] or "—",
            row["website"],
            row["wrong"],
            row["industry"],
            row["priority"],
            row["locality"],
            row["offer"],
            row.get("email_subject") or row.get("subject") or "",
            row["outreach"],
            "YES" if row["primary_email"] else "NO",
            "YES" if row["primary_phone"] else "NO",
        ]
        for c, val in enumerate(values, 4):
            cell = ws.cell(r, c, val)
            cell.alignment = wrap
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            if c == 14:  # Priority
                if "P1" in row["priority"]:
                    cell.fill = p1_fill
                elif "P2" in row["priority"]:
                    cell.fill = p2_fill
                else:
                    cell.fill = p3_fill

        for c in range(1, 21):
            ws.cell(r, c).border = thin

        ws.row_dimensions[r].height = 96

    widths = [10, 14, 16, 5, 30, 26, 28, 16, 32, 28, 36, 40, 28, 22, 24, 32, 42, 55, 10, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 36
    ws.auto_filter.ref = f"A1:T{len(rows)+1}"
    ws.freeze_panes = "E2"

    # ---------- Selected helper ----------
    ws_sel = wb.create_sheet("Selected Helper")
    ws_sel["A1"] = "Selected clients (YES in Contactable Prospects)"
    ws_sel["A1"].font = Font(bold=True, size=14)
    ws_sel["A3"] = "Instructions: After marking YES on Contactable Prospects, use Outreach_Console.html for true multi-send. Excel alone cannot open many WhatsApp tabs safely from one formula."
    ws_sel["A5"] = "Quick counts"
    ws_sel["A6"] = "Total contactable"
    ws_sel["B6"] = len(rows)
    ws_sel["A7"] = "With Email"
    ws_sel["B7"] = sum(1 for x in rows if x["primary_email"])
    ws_sel["A8"] = "With WhatsApp mobile"
    ws_sel["B8"] = sum(1 for x in rows if x["primary_phone"])
    ws_sel["A9"] = "With both"
    ws_sel["B9"] = sum(1 for x in rows if x["primary_email"] and x["primary_phone"])
    ws_sel.column_dimensions["A"].width = 100
    ws_sel.column_dimensions["B"].width = 12

    # ---------- Removed ----------
    all_leads = load_leads()
    removed = [l for l in all_leads if not usable(l)]
    ws_rm = wb.create_sheet("Removed (no contact)")
    ws_rm["A1"] = "Removed — no usable email AND no usable phone"
    ws_rm["A1"].font = Font(bold=True, size=12)
    for c, h in enumerate(["Company", "Industry", "Priority", "Locality", "Reason"], 1):
        style_header(ws_rm.cell(2, c, h), header_fill, header_font)
    for i, lead in enumerate(removed, 3):
        ws_rm.cell(i, 1, lead["company"])
        ws_rm.cell(i, 2, lead["industry"])
        ws_rm.cell(i, 3, lead["priority"])
        ws_rm.cell(i, 4, lead["locality"])
        ws_rm.cell(i, 5, "Email & phone both unavailable / directory-gated")
    for i, w in enumerate([34, 32, 24, 28, 48], 1):
        ws_rm.column_dimensions[get_column_letter(i)].width = w

    # ---------- Priority summary ----------
    ws2 = wb.create_sheet("By Priority")
    ws2["A1"] = "Priority mix (contactable only)"
    ws2["A1"].font = Font(bold=True, size=14)
    summary = [
        ("Priority", "Count"),
        ("P1 — Hot", sum(1 for l in rows if "P1" in l["priority"])),
        ("P2 — Redesign", sum(1 for l in rows if "P2" in l["priority"])),
        ("P3 — Upsell", sum(1 for l in rows if "P3" in l["priority"])),
        ("TOTAL", len(rows)),
    ]
    for r, row in enumerate(summary, 3):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(r, c, val)
            if r == 3:
                style_header(cell, header_fill, header_font)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 10

    wb.save(XLSX_OUT)


def build_html(rows: list[dict]) -> None:
    payload = []
    for r in rows:
        payload.append(
            {
                "id": r["source_id"],
                "company": r["company"],
                "owner": r["owner"],
                "email": r["primary_email"],
                "emails": r["emails"],
                "phone": r["primary_phone"],
                "phoneDisplay": r["primary_phone_display"],
                "industry": r["industry"],
                "priority": r["priority"],
                "locality": r["locality"],
                "offer": r["offer"],
                "wrong": r["wrong"],
                "message": r["outreach"],
                "subject": r["email_subject"],
                "mailto": r["mailto"],
                "whatsapp": r["whatsapp"],
            }
        )
    template = (ROOT / "outreach_console_template.html").read_text(encoding="utf-8")
    data_json = json.dumps(payload, ensure_ascii=False)
    if "__DATA_JSON__" not in template:
        raise RuntimeError("Template missing __DATA_JSON__ placeholder")
    HTML_OUT.write_text(template.replace("__DATA_JSON__", data_json), encoding="utf-8")



def main() -> None:
    all_leads = load_leads()
    rows = build_contactable(all_leads)
    removed = len(all_leads) - len(rows)

    JSON_OUT.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    build_xlsx(rows)
    build_html(rows)

    print(f"Original: {len(all_leads)}")
    print(f"Contactable kept: {len(rows)}")
    print(f"Removed (no email AND no phone): {removed}")
    print(f"With email: {sum(1 for r in rows if r['primary_email'])}")
    print(f"With WhatsApp mobile: {sum(1 for r in rows if r['primary_phone'])}")
    print(f"Excel: {XLSX_OUT}")
    print(f"Console: {HTML_OUT}")


if __name__ == "__main__":
    main()
