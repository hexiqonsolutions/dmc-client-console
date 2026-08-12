#!/usr/bin/env python3
"""
Build a quality-gated Navi Mumbai / MMR lead list.

Quality gates (strict):
- Must have a real 10-digit phone
- Must have website URL OR explicit no-website with Google-visible phone
- Must NOT reuse phones already in exclude list
- Must have individual website_issues (not generic template text)
- Prefer named owner

Usage:
  1. Drop researched batches into genuine_batches/*.json (arrays)
  2. python build_genuine_nm_leads.py
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
BATCH_DIR = ROOT / "genuine_batches"
OUT_JSON = ROOT / "genuine_nm_100.json"
OUT_XLSX = ROOT / "Genuine_NM_100_Clients.xlsx"
EXCLUDE_PHONES = set(json.loads((ROOT / "_exclude_phones.json").read_text(encoding="utf-8"))) if (ROOT / "_exclude_phones.json").exists() else set()
EXCLUDE_COMPANIES = set(json.loads((ROOT / "_exclude_companies.json").read_text(encoding="utf-8"))) if (ROOT / "_exclude_companies.json").exists() else set()

PHONE_RE = re.compile(r"(?:\+?91[\s\-]?)?([6-9]\d{9})")


def norm_phone(raw: str) -> str:
    digits = re.sub(r"\D", "", raw or "")
    if digits.startswith("91") and len(digits) >= 12:
        digits = digits[-10:]
    if len(digits) == 10 and digits[0] in "6789":
        return digits
    return ""


def display_phone(p: str) -> str:
    return f"+91 {p[:5]} {p[5:]}" if len(p) == 10 else p


def load_batches() -> list[dict]:
    rows: list[dict] = []
    if not BATCH_DIR.exists():
        return rows
    for path in sorted(BATCH_DIR.glob("*.json")):
        data = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(data, dict):
            data = data.get("leads") or data.get("prospects") or []
        for row in data:
            row["_batch"] = path.name
            rows.append(row)
    return rows


GENERIC_ISSUES = {
    "outdated website",
    "needs website",
    "website needs improvement",
    "looks outdated",
}


def is_specific_issue(text: str) -> bool:
    t = (text or "").strip().lower()
    if len(t) < 40:
        return False
    if t in GENERIC_ISSUES:
        return False
    # Require at least one concrete signal
    signals = [
        "booking", "whatsapp", "mobile", "http", "gmail", "copyright",
        "template", "flash", "jquery", "wordpress", "elementor", "order",
        "menu", "enquiry", "contact", "broken", "lorem", "table", "frame",
        "no website", "zomato", "swiggy", "justdial", "indiamart", "practo",
        "ssl", "slow", "popup", "stock photo", "no price", "no cta",
    ]
    return any(s in t for s in signals)


def qualify(row: dict) -> tuple[bool, str]:
    company = (row.get("company") or "").strip()
    if not company:
        return False, "missing company"
    if company.lower() in EXCLUDE_COMPANIES:
        return False, "already in old list"

    phone = norm_phone(str(row.get("phone") or ""))
    if not phone:
        # try phones list
        for p in row.get("phones") or []:
            phone = norm_phone(str(p))
            if phone:
                break
    if not phone:
        return False, "no real phone"
    # reject obviously fake / placeholder mobiles
    if len(set(phone)) <= 2 or phone in {"9999999999", "8888888888", "7777777777", "6666666666"}:
        return False, "fake phone"
    if phone.startswith(("000", "111", "12345")):
        return False, "fake phone"

    if phone in EXCLUDE_PHONES:
        return False, "phone already contacted/old list"

    issues = row.get("website_issues") or row.get("wrong") or ""
    if isinstance(issues, list):
        issues = "; ".join(issues)
    if not is_specific_issue(str(issues)):
        return False, "website issues too generic"

    website = (row.get("website") or "").strip()
    if not website:
        return False, "missing website field"
    # allow explicit no-website leads when issues document marketplace-only presence
    if website.lower() in {"no website", "none", "n/a"} and "no website" not in str(issues).lower() and "zomato" not in str(issues).lower():
        return False, "no website without clear no-website issue"

    owner = (row.get("owner") or row.get("owner_or_dm") or "").strip()
    row["_phone"] = phone
    row["_owner"] = owner
    row["_issues"] = str(issues).strip()
    return True, "ok"


def industry_bucket(industry: str) -> str:
    ind = (industry or "").lower()
    if "dental" in ind:
        return "dental"
    if any(x in ind for x in ("hospital", "maternity", "diagnostic", "physio", "dermat", "ent", "ophthal", "ortho", "ivf", "paediat", "health", "pet clinic", "nursing")):
        return "medical"
    if any(x in ind for x in ("banquet", "hotel", "guest", "wedding")):
        return "hospitality"
    if any(x in ind for x in ("bakery", "cafe", "restaurant", "f&b", "cater")):
        return "fnb"
    if any(x in ind for x in ("salon", "spa", "beauty")):
        return "salon"
    if any(x in ind for x in ("gym", "fitness", "yoga")):
        return "fitness"
    if "interior" in ind:
        return "interior"
    if any(x in ind for x in ("coach", "education", "tutor", "tuition")):
        return "education"
    if any(x in ind for x in ("packer", "travel", "real estate", "broker")):
        return "services"
    return "other"


# Soft caps so the published list is conversion-diverse (not dental-heavy).
INDUSTRY_CAPS = {
    "dental": 10,
    "interior": 12,
    "education": 8,
    "services": 10,
    "other": 12,
    "fitness": 8,
    "salon": 10,
    "fnb": 14,
    "hospitality": 14,
    "medical": 70,  # prioritize medical / specialty clinics
}


def score(row: dict) -> int:
    """Higher score = better chance of client conversion for outreach."""
    s = 0
    owner = (row.get("_owner") or "").strip()
    if owner and "not found" not in owner.lower():
        s += 35
        if owner.lower().startswith("dr") or "dr." in owner.lower():
            s += 8  # medical decision-makers often reply faster
    email = (row.get("email") or "").strip()
    if "@" in email and "not found" not in email.lower():
        s += 18
        if not any(x in email.lower() for x in ("gmail.", "yahoo.", "hotmail.", "rediff.")):
            s += 6  # branded email usually means more serious buyer
    if row.get("website", "").startswith("http"):
        s += 15

    bucket = industry_bucket(row.get("industry") or "")
    # Conversion priority: medical + hospitality/F&B/salon over dental
    bucket_pts = {
        "medical": 40,
        "hospitality": 28,
        "fnb": 26,
        "salon": 22,
        "fitness": 16,
        "interior": 14,
        "education": 12,
        "services": 10,
        "other": 8,
        "dental": -25,  # demote — list was oversaturated
    }
    s += bucket_pts.get(bucket, 0)

    issues = row.get("_issues", "").lower()
    for k, pts in [
        ("no booking", 22),
        ("online booking", 16),
        ("no online order", 18),
        ("order", 10),
        ("whatsapp", 10),
        ("gmail", 12),
        ("http", 12),
        ("500", 18),
        ("hello world", 20),
        ("lorem", 18),
        ("stuck at 0", 16),
        ("0+", 12),
        ("typo", 12),
        ("template", 12),
        ("copyright", 8),
        ("practo", 14),
        ("zomato", 12),
        ("swiggy", 12),
        ("placeholder", 16),
        ("broken", 14),
        ("mobile", 8),
    ]:
        if k in issues:
            s += pts
    s += min(len(row.get("_issues", "")) // 18, 18)
    return s


def select_with_caps(qualified: list[dict], target: int = 140) -> list[dict]:
    """Pick highest-conversion leads while respecting industry diversity caps."""
    qualified = sorted(qualified, key=lambda r: (-r["score"], r["company"].lower()))
    selected: list[dict] = []
    counts: Counter = Counter()
    for row in qualified:
        bucket = industry_bucket(row.get("industry") or "")
        cap = INDUSTRY_CAPS.get(bucket, 12)
        if counts[bucket] >= cap:
            continue
        selected.append(row)
        counts[bucket] += 1
        if len(selected) >= target:
            break
    return selected


def first_name(owner: str) -> str:
    if not owner or "not found" in owner.lower():
        return ""
    cleaned = re.sub(r"^(Dr\.?|Adv\.?|CA|Mr\.?|Mrs\.?|Ms\.?)\s+", "", owner.strip(), flags=re.I)
    part = re.split(r"[\s,/]", cleaned)[0]
    if len(part) < 2:
        return ""
    # Skip initials like R.R. / A.K.
    if re.fullmatch(r"[A-Za-z]\.?[A-Za-z]\.?", part) or (part.count(".") >= 1 and len(part) <= 5):
        return ""
    return part


def site_label(row: dict) -> str:
    w = (row.get("website") or "").strip()
    if not w.startswith("http"):
        return ""
    w = re.sub(r"/\(S\([^)]+\)\)", "", w)
    m = re.match(r"(https?://[^/\s]+)", w, re.I)
    if not m:
        return w
    host = m.group(1).rstrip("/") + "/"
    path = w[len(m.group(1)) :]
    if not path or path in {"/", "/index.html", "/index.php"} or len(path) > 40 or "ExternalSite" in path or "aspx" in path.lower():
        return host
    return w.split("?")[0].rstrip("/") + ("/" if w.endswith("/") else "")


def split_issues(text: str) -> list[str]:
    parts = re.split(r"[;•\n]+", text or "")
    return [p.strip(" -•").strip() for p in parts if p.strip()]


def plain_issue(raw: str, industry: str = "") -> str:
    """Turn internal audit notes into polite, specific observations."""
    t = (raw or "").strip()
    low = t.lower()
    ind = (industry or "").lower()

    if "no owned website" in low or low.startswith("no website") or "only zomato" in low or "only swiggy" in low:
        return "Most people currently find you on listings (Google / Zomato / Swiggy) rather than a site of your own"
    if "http 500" in low or "500 error" in low or "intermittently errors" in low or "returns server error" in low or "intermittently 500" in low or "returns 500" in low:
        return "On one visit the homepage showed an error instead of opening normally (this can happen if the server is overloaded)"
    if "lorem ipsum" in low:
        return "A little placeholder text is still visible — easy to replace with your own copy"
    if "hello world" in low:
        return "A default WordPress sample post is still live; replacing it with a short update would feel more current"
    if "stuck at 0" in low or "counters stuck" in low or "0+" in low or "0 locations" in low:
        return "A few numbers on the homepage still show 0 — visitors may think the page is still being set up"
    if "copyright" in low and ("2020" in low or "2021" in low or "2022" in low or "outdated" in low or "frozen" in low):
        return "The copyright year looks a little old, so some visitors may assume the site hasn't been updated recently"
    if "http-only" in low or "still on http" in low or "no https" in low:
        return "The site still opens without the browser lock (HTTPS). Adding it usually helps people feel safer sharing details"
    if "@gmail" in low or "gmail-only" in low or "gmail for business" in low or "gmail as" in low or "gmail contact" in low or (
        "gmail" in low and ("email" in low or "contact" in low)
    ):
        return "The public email is a Gmail address — a simple branded email (like info@yourbrand.in) often feels more professional"
    if "yahoo" in low:
        return "The public email is a Yahoo address — a branded company email usually feels more current"
    if "needhelp@floens" in low or "fake support" in low or ("placeholder" in low and "number" in low):
        return "A leftover template contact still appears on the site (not your real number/email) — worth swapping for yours"
    if "2012" in low and ("offer" in low or "valid" in low or "march" in low):
        return "An older offer is still showing (one still says valid till 2012) — updating it would make the page feel current"
    if "lorem" in low or "mirth large" in low or "unfinished" in low or "[location]" in low:
        return "A small bit of unfinished / sample text is still visible — easy to tidy up"
    if "encoding" in low or "mojibake" in low or "glitch" in low:
        return "A few reviews show broken characters; a quick text fix would make them easier to read"
    if "instagram" in low or "facebook" in low and ("broken" in low or "token" in low or "#0" in low):
        return "The Instagram / Facebook links don't currently open your real pages — a small fix if you'd like people to follow you"
    if "phone not" in low or "homepage lacks prominent phone" in low or "phone not on homepage" in low:
        return "The phone number isn't easy to spot on the first screen — putting it in the header usually helps"
    if "no public email" in low or "no visible business email" in low:
        return "There isn't a visible business email yet, so some people may hesitate to write in"
    if "elementor" in low and ("watermark" in low or "footer" in low):
        return "The footer still shows a website-builder credit — removing it is a small polish"
    if "viewport" in low or "weak mobile" in low or "mobile ux" in low:
        return "On a phone the layout is a little tight — making it mobile-friendly usually helps a lot in Navi Mumbai"
    if "wix" in low:
        return "The site still has a simple builder look; a cleaner custom layout would better match your brand"
    if "swiggy" in low or "zomato" in low:
        return "Ordering currently points people to Swiggy/Zomato — a direct order option on your own site can keep more of the margin"
    if "practo" in low:
        return "Appointments currently go through Practo — a simple booking option on your own site can capture more patients directly"
    if "template leftovers" in low or "fit365" in low or "luxe haven" in low or "solox" in low or "info@mysite.com" in low:
        return "A little leftover template branding is still showing — replacing it with your own name/logo is a quick win"
    if "no whatsapp" in low or "no clear online booking" in low or "no online booking" in low or "call-only" in low or "phone-only" in low or "phone-first" in low:
        if any(x in ind for x in ("dental", "clinic", "hospital", "health", "physio", "salon", "spa")):
            return "A clear Book Appointment button on the homepage would make it easier for patients (many currently have to look for a number)"
        if any(x in ind for x in ("hotel", "banquet", "guest")):
            return "A simple Check Availability / Book button would help guests book with you instead of going to Booking.com"
        if any(x in ind for x in ("bakery", "cafe", "restaurant", "f&b")):
            return "An Order Online or Book a table button on the homepage would make it easier for people who are ready to buy"
        return "A clear Enquire / Book button on the homepage would make it easier for people who don't want to call first"
    if "no online order" in low or "no cart" in low or "thin menu" in low:
        return "There's no easy online order yet — a simple menu/checkout often keeps people from switching to Swiggy/Zomato"
    if "ota" in low or "direct-booking" in low or "direct booking" in low:
        return "Bookings still go through listing sites — a short form on your own page can help guests book with you directly"
    if "google sites" in low:
        return "The current page is a basic Google Sites setup; a dedicated site would better match the quality of your work"
    if "typo" in low:
        m = re.search(r"['\"]([^'\"]{3,40})['\"]", t)
        if m:
            return f"A couple of small spelling slips on the homepage (e.g. '{m.group(1)}') — easy to fix, and it helps first impressions"
        return "A couple of small spelling slips on the homepage — easy to fix, and it helps first impressions"
    if "broken" in low and ("link" in low or "icon" in low or "social" in low or "#0" in low):
        return "A couple of buttons/links don't open a page yet — worth pointing them to the right place"
    if "template" in low or "elementor" in low or "wix" in low or "brochure" in low:
        if any(x in ind for x in ("dental", "clinic", "hospital")):
            return "The layout still feels like a standard clinic template — a more personal look usually helps patients choose you"
        if any(x in ind for x in ("interior",)):
            return "The pages feel a little generic; a stronger project gallery would better show your actual work"
        return "The layout still feels a little template-like — a more personal look would better reflect your brand"
    if "thin" in low or "dated" in low or "old html" in low or "php" in low:
        return "The pages are a little light on detail — a bit more proof (photos, services, reviews) usually helps people feel ready to call"
    cleaned = re.sub(r"\s+", " ", t).strip()
    if len(cleaned) > 140:
        cleaned = cleaned[:137].rsplit(" ", 1)[0] + "…"
    return cleaned[0].upper() + cleaned[1:] if cleaned else t


def industry_why(industry: str) -> str:
    ind = (industry or "").lower()
    if "dental" in ind:
        return "Patients usually open 2–3 dentist sites on their phone and book the one that feels easiest. A clearer booking path often helps."
    if any(x in ind for x in ("ivf", "fertility")):
        return "Fertility searches are high-intent. Couples usually shortlist clinics whose site feels clear and trustworthy before they call."
    if any(x in ind for x in ("maternity", "gynae", "woman", "nursing")):
        return "Expecting parents compare maternity options carefully online. Clear packages and easy booking often tip the decision."
    if any(x in ind for x in ("diagnostic", "patholog", "lab", "imaging")):
        return "People usually book labs from their phone for packages and home collection. A clearer booking path often wins that booking."
    if any(x in ind for x in ("physio",)):
        return "Patients looking for physio often book the clinic that makes the next session easiest to schedule."
    if any(x in ind for x in ("dermat", "skin", "aesthetic")):
        return "Skin and aesthetic consults are often decided after a quick look at treatment pages and how easy booking feels."
    if any(x in ind for x in ("ent", "eye", "ophthal", "ortho", "pet")):
        return "Specialty patients usually open 2–3 clinic sites and choose the one that feels clearest to book."
    if any(x in ind for x in ("hospital", "health", "ivf", "diagnostic", "maternity", "dermat")):
        return "Most patients compare clinics on their phone before they call. A clearer booking page often makes that choice easier."
    if any(x in ind for x in ("bakery", "cafe", "restaurant", "f&b", "cater")):
        return "When someone is ready to order, they usually pick whichever option feels simplest. A clearer order/book path on your own site can help."
    if any(x in ind for x in ("hotel", "banquet", "guest", "wedding")):
        return "Guests who land on your site often want to check dates quickly. Making that easy can bring more direct bookings."
    if "interior" in ind:
        return "Homeowners usually shortlist 2–3 designers from Google. A clear project gallery and quote form often wins the site visit."
    if any(x in ind for x in ("gym", "fitness", "yoga", "salon", "spa", "beauty")):
        return "People often decide trial/memberships from their phone. A simple booking button usually makes that easier."
    if any(x in ind for x in ("coach", "education", "tutor")):
        return "Parents compare institutes online first. A simple demo-class / enquiry form often captures the interest while it's warm."
    if any(x in ind for x in ("packers", "travel", "real estate", "broker")):
        return "For this kind of work, people like a clear, trustworthy site with an easy way to ask for a quote."
    return "People who find you on Google usually decide in a few seconds whether to enquire. A clearer next step on the site often helps."


def client_points(row: dict, limit: int = 3) -> list[str]:
    industry = row.get("industry") or ""
    seen = set()
    out = []
    for raw in split_issues(row.get("_issues") or ""):
        p = plain_issue(raw, industry)
        key = p.lower()
        if key in seen or len(p) < 20:
            continue
        seen.add(key)
        out.append(p)
        if len(out) >= limit:
            break
    return out or [plain_issue(row.get("_issues") or "", industry)]


def draft_whatsapp(row: dict) -> str:
    first = first_name(row.get("_owner") or "")
    greet = f"Hi {first}," if first else "Hello,"
    company = row["company"]
    url = site_label(row)
    points = client_points(row, 2)
    bullets = "\n".join(f"• {p}" for p in points)
    why = industry_why(row.get("industry") or "")
    if url:
        opened = (
            f"I came across {company} and had a look at {url}. "
            f"Hope you don't mind me sharing a couple of small observations — only if useful:\n{bullets}"
        )
    else:
        opened = (
            f"I came across {company} while looking at local businesses. "
            f"Hope you don't mind me sharing a couple of small observations — only if useful:\n{bullets}"
        )
    return (
        f"{greet}\n\n"
        f"I'm Vaibhav from DMC Creatives Studio. {opened}\n\n"
        f"{why}\n\n"
        f"If it would help, I can send a free one-page concept for {company} this week — no charge and no obligation, just so you can see the idea.\n\n"
        f"Vaibhav Gurav\nDMC Creatives Studio\nwww.dmcstudio.in\n+91 83693 61785"
    )


def draft_email(row: dict) -> tuple[str, str]:
    first = first_name(row.get("_owner") or "")
    greet = f"Dear {first}," if first else "Hello,"
    company = row["company"]
    url = site_label(row)
    locality = row.get("locality") or "Navi Mumbai"
    points = client_points(row, 3)
    bullets = "\n".join(f"• {p}" for p in points)
    why = industry_why(row.get("industry") or "")
    host = url.replace("https://", "").replace("http://", "").rstrip("/") if url else ""
    if url:
        subject = f"A quick, polite note on {host}"
        opened = (
            f"I hope you're well. I'm Vaibhav from DMC Creatives Studio. "
            f"While looking at {industry_label_short(row.get('industry') or 'local')} businesses in {locality}, "
            f"I visited {url}."
        )
    else:
        subject = f"A quick note for {company}"
        opened = (
            f"I hope you're well. I'm Vaibhav from DMC Creatives Studio. "
            f"I came across {company} in {locality} — most of the online presence right now is through listings rather than a site of your own."
        )
    body = (
        f"{greet}\n\n"
        f"{opened}\n\n"
        f"I wanted to share a few small observations, only in case they're helpful:\n{bullets}\n\n"
        f"{why}\n\n"
        f"If you'd like, I can share a free concept page for {company} this week so you can see what a clearer version might look like. No obligation at all.\n\n"
        f"Warm regards,\nVaibhav Gurav\nDMC Creatives Studio\nhello@dmcstudio.in\nwww.dmcstudio.in\n+91 83693 61785"
    )
    return subject, body


def industry_label_short(industry: str) -> str:
    return (industry or "local").split("/")[0].strip().lower() or "local"


def main() -> None:
    BATCH_DIR.mkdir(exist_ok=True)
    raw = load_batches()
    print(f"Loaded {len(raw)} raw rows from batches")

    qualified: list[dict] = []
    rejected = Counter()
    seen_phones: set[str] = set()
    seen_companies: set[str] = set()

    for row in raw:
        ok, reason = qualify(row)
        if not ok:
            rejected[reason] += 1
            continue
        phone = row["_phone"]
        company_key = row["company"].strip().lower()
        if phone in seen_phones:
            rejected["duplicate phone in new set"] += 1
            continue
        if company_key in seen_companies:
            rejected["duplicate company in new set"] += 1
            continue
        seen_phones.add(phone)
        seen_companies.add(company_key)
        row["score"] = score(row)
        qualified.append(row)

    qualified.sort(key=lambda r: (-r["score"], r["company"].lower()))
    print(f"Qualified before caps: {len(qualified)}")
    print("Rejected:", dict(rejected))
    by_bucket = Counter(industry_bucket(r.get("industry") or "") for r in qualified)
    print("Qualified buckets:", dict(by_bucket))

    selected = select_with_caps(qualified, target=140)
    print(f"Selected after diversity caps: {len(selected)}")
    print("Selected buckets:", dict(Counter(industry_bucket(r.get("industry") or "") for r in selected)))

    final = []
    for i, row in enumerate(selected, 1):
        subject, email_body = draft_email(row)
        wa = draft_whatsapp(row)
        phone = row["_phone"]
        item = {
            "id": i,
            "company": row["company"].strip(),
            "owner": row.get("_owner") or "",
            "phone": display_phone(phone),
            "phone_digits": "91" + phone,
            "email": (row.get("email") or "").strip(),
            "website": (row.get("website") or "").strip(),
            "locality": (row.get("locality") or "").strip(),
            "industry": (row.get("industry") or "").strip(),
            "website_issues": row["_issues"],
            "why_buy": (row.get("why_buy") or row.get("offer") or "").strip(),
            "source": (row.get("source") or row.get("_batch") or "").strip(),
            "score": row["score"],
            "subject": subject,
            "email_body": email_body,
            "whatsapp_msg": wa,
            "wa_link": f"https://wa.me/91{phone}?text=" + __import__("urllib.parse").parse.quote(wa),
            "mailto": (
                f"mailto:{row.get('email')}?subject={__import__('urllib.parse').parse.quote(subject)}"
                f"&body={__import__('urllib.parse').parse.quote(email_body)}"
                if row.get("email") and "@" in row.get("email", "")
                else ""
            ),
            "batch": row.get("_batch", ""),
        }
        final.append(item)

    OUT_JSON.write_text(json.dumps(final, indent=2, ensure_ascii=False), encoding="utf-8")
    write_xlsx(final)
    print(f"Wrote {len(final)} leads -> {OUT_JSON.name}, {OUT_XLSX.name}")
    with_owner = sum(1 for x in final if x["owner"])
    with_email = sum(1 for x in final if x["email"] and "@" in x["email"])
    print(f"With owner name: {with_owner}/{len(final)} | With email: {with_email}/{len(final)}")


def write_xlsx(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Genuine Clients"
    headers = [
        "ID", "Company", "Owner", "Phone", "Email", "Website", "Locality",
        "Industry", "Website Issues (specific)", "Why they'll buy", "Score",
        "Email Subject", "Email Body", "WhatsApp Message",
        "WhatsApp Link", "Email Link", "Source",
    ]
    header_fill = PatternFill("solid", fgColor="1F3D2B")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for r, row in enumerate(rows, 2):
        vals = [
            row["id"], row["company"], row["owner"], row["phone"], row["email"],
            row["website"], row["locality"], row["industry"], row["website_issues"],
            row["why_buy"], row["score"], row.get("subject") or "", row.get("email_body") or "",
            row.get("whatsapp_msg") or "", row["wa_link"], row["mailto"], row["source"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [5, 28, 22, 16, 28, 34, 22, 16, 48, 36, 8, 36, 48, 48, 18, 18, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{len(rows)+1}"
    wb.save(OUT_XLSX)


if __name__ == "__main__":
    main()
