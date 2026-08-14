#!/usr/bin/env python3
"""Score researched website-need prospects, write Excel + WhatsApp drafts.

Only uses phones/emails/owners supplied in research JSON. Never invents contacts.
Public listed mobiles are treated as business WhatsApp (wa.me) — labelled as such.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
RAW = ROOT / "website_need_research.json"
OUT_JSON = ROOT / "website_need_100.json"
OUT_XLSX = ROOT / "Website_Need_Prospects_100.xlsx"
OUT_HTML = ROOT / "Website_Need_Outreach.html"

EXCLUDE_PHONES = set()
if (ROOT / "_exclude_phones.json").exists():
    EXCLUDE_PHONES = set(json.loads((ROOT / "_exclude_phones.json").read_text(encoding="utf-8")))

PHONE_RE = re.compile(r"(?:\+?91[\s\-]?)?([6-9]\d{9})")


def digits_phone(raw) -> str:
    if raw is None:
        return ""
    m = PHONE_RE.search(str(raw).replace(" ", ""))
    if not m:
        d = re.sub(r"\D", "", str(raw))
        if d.startswith("91") and len(d) >= 12:
            d = d[-10:]
        if len(d) == 10 and d[0] in "6789":
            return d
        return ""
    return m.group(1)


def display_phone(p: str) -> str:
    return f"+91 {p[:5]} {p[5:]}" if len(p) == 10 else ""


def first_name(owner: str) -> str:
    if not owner or owner.lower() in {"not identified", "unknown", "n/a"}:
        return ""
    part = re.split(r"[/(,]", owner)[0].strip()
    part = re.sub(r"^(Dr\.?|CA|Mr\.?|Mrs\.?|Ms\.?)\s+", "", part, flags=re.I)
    tok = part.split()
    return tok[0] if tok else ""


def norm_status(s: str) -> str:
    t = (s or "").lower()
    if "social" in t:
        return "C — Social media only"
    if "directory" in t or t.startswith("d"):
        return "D — Directory only"
    if "no website" in t or t.startswith("a") or "no independent" in t or "no dedicated" in t:
        return "A — No website"
    if "poor" in t or "weak" in t or "broken" in t or "outdated" in t or t.startswith("b"):
        return "B — Poor website"
    return (s or "Unknown").strip() or "Unknown"


def score_row(r: dict) -> int:
    year = r.get("estimated_start_year")
    try:
        y = int(year) if year not in (None, "", "null") else None
    except (TypeError, ValueError):
        y = None
    newness = 5
    if y == 2026:
        newness = 20
    elif y == 2025:
        newness = 18
    elif y == 2024:
        newness = 15
    elif y == 2023:
        newness = 10
    elif y and y >= 2020:
        newness = 6

    st = r["website_status"]
    opp = 10
    if st.startswith("A"):
        opp = 30
    elif st.startswith("C"):
        opp = 28
    elif st.startswith("D"):
        opp = 25
    elif st.startswith("B"):
        opp = 20

    cat = (r.get("category") or "").lower()
    loc = (r.get("locality") or "").lower()
    growth = 8
    if y and y >= 2024:
        growth += 6
    if any(x in cat for x in ("interior", "architect", "fabricat", "manufact", "contractor", "waterproof")):
        growth += 4
    if "instagram" in (r.get("instagram") or ""):
        growth += 2
    growth = min(20, growth)

    roi = 8
    if any(x in cat for x in ("interior", "architect", "manufact", "cnc", "fabricat", "contractor", "real estate", "clinic", "coach")):
        roi = 18
    elif any(x in cat for x in ("gym", "salon", "cafe", "bakery", "wedding", "photo", "detail")):
        roi = 14
    elif "packag" in cat or "engineer" in cat:
        roi = 16
    roi = min(20, roi)

    contact = 0
    if r.get("phone"):
        contact += 6
    if r.get("email"):
        contact += 2
    if first_name(r.get("owner") or ""):
        contact += 2
    contact = min(10, contact)

    return min(100, newness + opp + growth + roi + contact)


def priority_label(score: int) -> str:
    if score >= 80:
        return "HOT"
    if score >= 65:
        return "WARM"
    if score >= 50:
        return "COLD"
    return "REJECT"


def project_value(r: dict) -> str:
    cat = (r.get("category") or "").lower()
    if any(x in cat for x in ("manufact", "engineer", "fabricat", "architect", "contractor", "waterproof", "pressure")):
        return "High (₹30,000–₹75,000+)"
    if any(x in cat for x in ("interior", "coach", "clinic", "recruit", "ca ", "tax")):
        return "High (₹30,000–₹75,000+)"
    if any(x in cat for x in ("gym", "salon", "wedding", "photo", "cafe", "bakery", "detail", "furniture")):
        return "Medium (₹15,000–₹30,000)"
    return "Medium (₹15,000–₹30,000)"


def wa_message(r: dict) -> str:
    name = first_name(r.get("owner") or "")
    hi = f"Hi {name}," if name else "Hi,"
    company = r["company"]
    loc = r.get("locality") or "Navi Mumbai"
    opp = (r.get("website_opportunity") or "").strip()
    bullet = opp.split(".")[0].strip()
    if bullet and not bullet.endswith("."):
        bullet += "."
    st = r["website_status"]
    if st.startswith("A") or st.startswith("C") or st.startswith("D"):
        gap = "I couldn't find a proper website of your own — most people seem to find you on Instagram, Google or directories."
    else:
        url = r.get("website_url") or "your current site"
        gap = f"I had a look at {url} — it works, but it doesn't yet convert visitors the way a clear enquiry site would."
    rec = r.get("recommended_website") or "a simple mobile-first website"
    return (
        f"{hi}\n\n"
        f"I'm Vaibhav from DMC Creatives Studio. I came across {company} in {loc}.\n\n"
        f"{gap}\n"
        f"{bullet}\n\n"
        f"If useful, I can share a free one-page concept ({rec.lower()}) this week — no charge, no obligation.\n\n"
        f"Vaibhav Gurav\nDMC Creatives Studio\nwww.dmcstudio.in\n+91 83693 61785"
    )


def main() -> None:
    raw = json.loads(RAW.read_text(encoding="utf-8"))
    seen_co = set()
    seen_ph = set()
    out = []
    skipped = []
    for row in raw:
        company = (row.get("company") or "").strip()
        if not company:
            continue
        key = re.sub(r"\W+", "", company.lower())
        if key in seen_co:
            skipped.append((company, "duplicate name"))
            continue
        phone = digits_phone(row.get("phone"))
        if phone and phone in EXCLUDE_PHONES:
            skipped.append((company, "already contacted"))
            continue
        if phone and phone in seen_ph:
            skipped.append((company, "duplicate phone"))
            continue
        owner = (row.get("owner") or "").strip()
        if owner.lower() in {"not identified", "unknown", "n/a"}:
            owner = ""
        status = norm_status(row.get("website_status") or "")
        rec = {
            "company": company,
            "category": row.get("category") or "",
            "locality": row.get("locality") or "",
            "estimated_start_year": row.get("estimated_start_year"),
            "start_year_confidence": row.get("start_year_confidence") or "Unknown",
            "website_status": status,
            "website_url": row.get("website_url") or "",
            "google_maps": row.get("google_maps") or "",
            "instagram": row.get("instagram") or "",
            "phone": phone,
            "phone_display": display_phone(phone),
            "whatsapp": display_phone(phone) if phone else "",
            "whatsapp_note": "Public listed business mobile — use as WhatsApp if they have it on this number. Not a private personal number." if phone else "No public mobile found — use Instagram DM / Google listing.",
            "email": (row.get("email") or "").strip(),
            "owner": owner or "Not identified",
            "evidence_newness": row.get("evidence_newness") or "",
            "website_opportunity": row.get("website_opportunity") or "",
            "recommended_website": row.get("recommended_website") or "",
            "source_links": row.get("source_links") or [],
            "confidence": row.get("confidence") or "",
        }
        rec["lead_score"] = score_row(rec)
        rec["priority"] = priority_label(rec["lead_score"])
        rec["potential_project_value"] = project_value(rec)
        rec["whatsapp_msg"] = wa_message(rec) if phone else ""
        rec["wa_link"] = (
            f"https://wa.me/91{phone}?text={quote(rec['whatsapp_msg'])}" if phone else ""
        )
        rec["best_reason"] = rec["website_opportunity"]
        if rec["priority"] == "REJECT":
            skipped.append((company, f"score {rec['lead_score']}"))
            continue
        seen_co.add(key)
        if phone:
            seen_ph.add(phone)
        out.append(rec)

    out.sort(key=lambda x: (-x["lead_score"], x["company"]))
    for i, r in enumerate(out, 1):
        r["id"] = i

    OUT_JSON.write_text(json.dumps(out, indent=2, ensure_ascii=False), encoding="utf-8")
    write_xlsx(out)
    write_html(out)
    print("qualified", len(out), "skipped", len(skipped))
    print("with_phone", sum(1 for r in out if r["phone"]))
    print("HOT", sum(1 for r in out if r["priority"] == "HOT"))
    print("WARM", sum(1 for r in out if r["priority"] == "WARM"))
    print("A", sum(1 for r in out if r["website_status"].startswith("A")))
    print("B", sum(1 for r in out if r["website_status"].startswith("B")))
    print("C", sum(1 for r in out if r["website_status"].startswith("C")))
    print("D", sum(1 for r in out if r["website_status"].startswith("D")))
    print("wrote", OUT_JSON.name, OUT_XLSX.name, OUT_HTML.name)


def write_xlsx(rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"
    headers = [
        "#", "Business Name", "Category", "Location", "Business Age", "Age confidence",
        "Website Status", "Website URL", "Google Maps", "Instagram", "Phone / WhatsApp",
        "WhatsApp note", "Email", "Decision Maker", "Evidence of Newness",
        "Website Opportunity", "Recommended Website", "Lead Score", "Priority",
        "Potential Project Value", "Best Reason to Contact", "WhatsApp draft", "WhatsApp link",
        "Source Links", "Confidence",
    ]
    head_fill = PatternFill("solid", fgColor="14352C")
    head_font = Font(color="FFFFFF", bold=True)
    hot = PatternFill("solid", fgColor="FDE68A")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for r in rows:
        age = r.get("estimated_start_year") or "Unknown"
        vals = [
            r["id"], r["company"], r["category"], r["locality"], age, r["start_year_confidence"],
            r["website_status"], r["website_url"], r["google_maps"], r["instagram"],
            r["phone_display"], r["whatsapp_note"], r["email"], r["owner"], r["evidence_newness"],
            r["website_opportunity"], r["recommended_website"], r["lead_score"], r["priority"],
            r["potential_project_value"], r["best_reason"], r["whatsapp_msg"], r["wa_link"],
            " | ".join(r["source_links"]), r["confidence"],
        ]
        ws.append(vals)
        if r["priority"] == "HOT":
            for c in range(1, len(headers) + 1):
                ws.cell(ws.max_row, c).fill = hot
        for c in range(1, len(headers) + 1):
            ws.cell(ws.max_row, c).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [6, 32, 28, 28, 12, 16, 22, 32, 28, 28, 18, 36, 28, 22, 40, 40, 36, 10, 10, 22, 40, 40, 28, 40, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28

    sm = wb.create_sheet("Summary", 0)
    sm["A1"] = "Website-need prospects — Mumbai / Navi Mumbai"
    sm["A1"].font = Font(bold=True, size=14, color="14352C")
    stats = [
        ("Qualified leads", len(rows)),
        ("With public phone / WhatsApp", sum(1 for r in rows if r["phone"])),
        ("Named decision maker", sum(1 for r in rows if r["owner"] != "Not identified")),
        ("HOT (80+)", sum(1 for r in rows if r["priority"] == "HOT")),
        ("WARM (65–79)", sum(1 for r in rows if r["priority"] == "WARM")),
        ("COLD (50–64)", sum(1 for r in rows if r["priority"] == "COLD")),
        ("A — No website", sum(1 for r in rows if r["website_status"].startswith("A"))),
        ("B — Poor website", sum(1 for r in rows if r["website_status"].startswith("B"))),
        ("C — Social only", sum(1 for r in rows if r["website_status"].startswith("C"))),
        ("D — Directory only", sum(1 for r in rows if r["website_status"].startswith("D"))),
    ]
    sm["A3"] = "Metric"
    sm["B3"] = "Count"
    for i, (k, v) in enumerate(stats, 4):
        sm.cell(i, 1, k)
        sm.cell(i, 2, v)
    sm["A16"] = "WhatsApp rule"
    sm["A17"] = "Only public business mobiles from listings/sites/Instagram bios. Not private personal numbers. Click the WhatsApp link in the Prospects sheet."
    sm.column_dimensions["A"].width = 42
    sm.column_dimensions["B"].width = 14
    wb.save(OUT_XLSX)


def write_html(rows: list[dict]) -> None:
    cards = []
    for r in rows:
        wa = f'<a class="btn" href="{r["wa_link"]}">WhatsApp</a>' if r["wa_link"] else '<span class="muted">No public WhatsApp</span>'
        ig = f'<a href="{r["instagram"]}">Instagram</a>' if r["instagram"] else ""
        web = f'<a href="{r["website_url"]}">{r["website_url"]}</a>' if r["website_url"] else "No site"
        owner = r["owner"]
        cards.append(
            f"""<article class="card {r['priority'].lower()}">
  <header><b>{r['id']}. {r['company']}</b> <span class="tag">{r['priority']}</span> <span class="score">{r['lead_score']}</span></header>
  <p class="meta">{r['category']} · {r['locality']} · {r['website_status']}</p>
  <p><b>Owner:</b> {owner} · <b>Phone:</b> {r['phone_display'] or 'Not public'}</p>
  <p>{r['website_opportunity']}</p>
  <div class="actions">{wa} {ig} {web}</div>
  <pre>{r['whatsapp_msg']}</pre>
</article>"""
        )
    html = f"""<!DOCTYPE html><html lang="en"><head><meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>Website-need prospects — DMC Creatives</title>
<style>
body{{font-family:system-ui,sans-serif;background:#F5F7F6;color:#14201C;margin:0}}
.wrap{{max-width:980px;margin:0 auto;padding:24px}}
h1{{color:#14352C}}
.card{{background:#fff;border:1px solid #E2EAE6;border-radius:16px;padding:16px;margin:12px 0}}
.card.hot{{border-color:#B8956A}}
.tag{{background:#14352C;color:#fff;border-radius:999px;padding:2px 8px;font-size:12px}}
.score{{float:right;font-weight:700;color:#B8956A}}
.meta,.muted{{color:#5A6B63;font-size:14px}}
.btn{{display:inline-block;background:#128C7E;color:#fff;padding:8px 12px;border-radius:10px;text-decoration:none;margin-right:8px}}
pre{{white-space:pre-wrap;background:#EEF3F0;padding:12px;border-radius:12px;font-size:13px}}
a{{color:#14352C}}
</style></head><body><div class="wrap">
<h1>Website-need prospects</h1>
<p>{len(rows)} verified MMR businesses. WhatsApp uses publicly listed business mobiles only.</p>
{''.join(cards)}
</div></body></html>"""
    OUT_HTML.write_text(html, encoding="utf-8")


if __name__ == "__main__":
    main()
