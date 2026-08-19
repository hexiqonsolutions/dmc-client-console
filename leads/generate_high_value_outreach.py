#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build Gulf (GCC) client list + BuildView-style one-click outreach console."""

from __future__ import annotations

import json
import re
import urllib.parse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:\Users\FPIN\Desktop\dm os\leads")
GULF = ROOT / "gulf_verified.json"
JSON_OUT = ROOT / "high_value_prospects.json"
XLSX_OUT = ROOT / "Gulf_Clients.xlsx"
HTML_OUT = ROOT / "DMC_Client_Console.html"

SIGN = (
    "Regards,\n"
    "Vaibhav Gurav\n"
    "DMC Creatives Studio\n"
    "hello@dmcstudio.in\n"
    "www.dmcstudio.in\n"
    "+91 83693 61785"
)

INDUSTRY_ANGLE = {
    "Healthcare": "patients compare clinics online before they call — a clear website and booking path wins the enquiry",
    "Hotel / F&B": "guests and corporates book the hotel that looks trustworthy and easy on mobile",
    "Manufacturing": "B2B buyers shortlist suppliers from Google and catalogue sites before they raise an RFQ",
    "Education": "parents and students decide admissions from your website long before they walk in",
    "Automotive": "test-drive and service bookings now start online — weak dealer sites lose leads to competitors",
    "Real Estate / Design": "homebuyers and project clients judge your brand from the first project page they open",
    "Fitness": "memberships convert when the studio looks current and booking is one tap away",
    "Services": "serious clients still Google you — a weak site quietly costs you deals",
    "Trading": "importers and distributors lose RFQs when the catalogue site looks older than the stock",
    "Business": "buyers in the Gulf judge readiness from your website before they ever call",
}


def extract_emails(raw: str) -> list[str]:
    if not raw:
        return []
    low = raw.lower()
    if "not found" in low or "gated" in low or "form only" in low:
        return []
    found = re.findall(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", raw)
    out = []
    for e in found:
        el = e.lower()
        if "…" in e or e.endswith("@"):
            continue
        if re.search(r"@(g|ymail|yahoo|rediffmail|hotmail|gmail|vsnl)\.?$", el):
            # allow full gmail/yahoo etc — only skip truncated
            pass
        if e not in out:
            out.append(e)
    return out


GCC_CC = {
    "971": "UAE",
    "966": "Saudi Arabia",
    "974": "Qatar",
    "965": "Kuwait",
    "973": "Bahrain",
    "968": "Oman",
}

LOCALITY_CC = (
    ("united arab", "971"),
    ("u.a.e", "971"),
    ("uae", "971"),
    ("dubai", "971"),
    ("abu dhabi", "971"),
    ("sharjah", "971"),
    ("ajman", "971"),
    ("fujairah", "971"),
    ("ras al khaimah", "971"),
    ("umm al quwain", "971"),
    ("saudi", "966"),
    ("riyadh", "966"),
    ("jeddah", "966"),
    ("dammam", "966"),
    ("khobar", "966"),
    ("qatar", "974"),
    ("doha", "974"),
    ("kuwait", "965"),
    ("bahrain", "973"),
    ("manama", "973"),
    ("sitra", "973"),
    ("oman", "968"),
    ("muscat", "968"),
    ("sohar", "968"),
    ("salalah", "968"),
    ("barka", "968"),
)

# Mobile prefixes after country code
GCC_MOBILE_RE = [
    ("971", re.compile(r"^9715\d{8}$")),
    ("966", re.compile(r"^9665\d{8}$")),
    ("974", re.compile(r"^974[3567]\d{7}$")),
    ("965", re.compile(r"^965[569]\d{7}$")),
    ("973", re.compile(r"^973[36]\d{7}$")),
    ("968", re.compile(r"^968[79]\d{7}$")),
]


def country_cc(locality: str) -> str:
    low = (locality or "").lower()
    for needle, cc in LOCALITY_CC:
        if needle in low:
            return cc
    return ""


def _digits(raw: str) -> str:
    d = re.sub(r"\D", "", raw or "")
    if d.startswith("00"):
        d = d[2:]
    return d


def extract_mobiles(raw: str, locality: str = "") -> list[str]:
    if not raw:
        return []
    low = raw.lower()
    if "not found" in low or "gated" in low or "show number" in low:
        return []
    found: list[str] = []

    def add(num: str) -> None:
        if not num or num in found:
            return
        for _, rx in GCC_MOBILE_RE:
            if rx.match(num):
                found.append(num)
                return

    digits = _digits(raw)
    for pat in (
        r"9715\d{8}",
        r"9665\d{8}",
        r"974[3567]\d{7}",
        r"965[569]\d{7}",
        r"973[36]\d{7}",
        r"968[79]\d{7}",
    ):
        for m in re.finditer(pat, digits):
            add(m.group(0))

    cc = country_cc(locality)
    for chunk in re.split(r"[,;/|]|WhatsApp|whatsapp|Tel|tel|Mob|mob", raw):
        d = _digits(chunk)
        if not d:
            continue
        if cc in {"971", "966"} and re.fullmatch(r"0?5\d{8}", d):
            add(cc + d.lstrip("0"))
        elif cc == "974" and re.fullmatch(r"0?[3567]\d{7}", d):
            add("974" + d.lstrip("0"))
        elif cc == "965" and re.fullmatch(r"0?[569]\d{7}", d):
            add("965" + d.lstrip("0"))
        elif cc == "973" and re.fullmatch(r"0?[36]\d{7}", d):
            add("973" + d.lstrip("0"))
        elif cc == "968" and re.fullmatch(r"0?[79]\d{7}", d):
            add("968" + d.lstrip("0"))
        else:
            add(d)
    return found


def has_any_phone(raw: str) -> bool:
    if not raw:
        return False
    low = raw.lower()
    if "not found" in low or "gated" in low or "show number" in low:
        return False
    digits = _digits(raw)
    return len(digits) >= 8


def display_phone(e164: str) -> str:
    if e164.startswith("971") and len(e164) == 12:
        return f"+971 {e164[3:5]} {e164[5:8]} {e164[8:]}"
    if e164.startswith("966") and len(e164) == 12:
        return f"+966 {e164[3:5]} {e164[5:8]} {e164[8:]}"
    if e164.startswith("974") and len(e164) == 11:
        return f"+974 {e164[3:7]} {e164[7:]}"
    if e164.startswith("965") and len(e164) == 11:
        return f"+965 {e164[3:7]} {e164[7:]}"
    if e164.startswith("973") and len(e164) == 11:
        return f"+973 {e164[3:7]} {e164[7:]}"
    if e164.startswith("968") and len(e164) == 11:
        return f"+968 {e164[3:7]} {e164[7:]}"
    if e164.startswith("91") and len(e164) == 12:
        return "+91 " + e164[2:7] + " " + e164[7:]
    return "+" + e164 if e164 and not e164.startswith("+") else e164


def first_name(dm: str) -> str:
    if not dm:
        return ""
    dm = re.sub(r"\([^)]*\)", "", dm).strip()
    parts = dm.replace(",", " ").replace("/", " ").split()
    skip = {"dr", "dr.", "mr", "mr.", "mrs", "mrs.", "ms", "ms.", "and", "team"}
    for p in parts:
        if p.lower().rstrip(".") in skip:
            continue
        if p and p[0].isalpha() and len(p) > 1:
            return p
    return ""


def greeting(dm: str) -> str:
    fn = first_name(dm)
    return f"Hi {fn}," if fn else "Hello,"


def subject_line(company: str, industry: str, wrong: str) -> str:
    short = company.split("(")[0].strip()
    w = (wrong or "").lower()
    if "no website" in w or "directory" in w or "instagram" in w or "coming soon" in w or "no owned" in w:
        return f"{short}: buyers search Google before they call"
    if "new" in w and ("business" in w or "launch" in w or "2024" in w or "2025" in w or "2026" in w):
        return f"{short}: a new Gulf brand still needs a real website"
    if "http" in w and "https" not in w:
        return f"{short} is still on HTTP — that hurts trust"
    if industry == "Healthcare":
        return f"{short}: patients shortlist clinics online first"
    if industry == "Hotel / F&B":
        return f"{short}: your website vs today's direct bookings"
    if industry == "Manufacturing":
        return f"{short}: RFQ buyers want a real catalogue site"
    if industry == "Education":
        return f"{short}: admissions start on your website"
    if industry == "Automotive":
        return f"{short}: test-drive leads are leaking online"
    if industry == "Real Estate / Design":
        return f"{short}: project pages that convert enquiries"
    return f"{short}'s website still looks stuck in the past"


def email_body(r: dict) -> str:
    company = r["company"]
    industry = r["industry"]
    locality = r["locality"]
    wrong = r["wrong"]
    offer = r["offer"]
    g = greeting(r.get("owner_or_dm") or "")
    angle = INDUSTRY_ANGLE.get(industry, INDUSTRY_ANGLE["Business"])
    area = locality or "the GCC"
    lead_type = (r.get("lead_type") or "").lower()
    open_line = (
        f"While reviewing {industry.lower()} businesses in {area}, I came across {company}."
    )
    wrong_l = wrong.lower()
    if lead_type == "new_business" or "no website" in wrong_l or "coming soon" in wrong_l or "instagram" in wrong_l:
        problem = (
            f"{company} looks like a new / growing Gulf brand, but buyers still struggle "
            "to find a strong owned website they can trust."
        )
    elif "no website" in wrong_l or "directory" in wrong_l:
        problem = f"{company} still lacks a strong owned website buyers can trust."
    else:
        problem = (
            f"I reviewed {company}'s current digital presence — it works, "
            "but it no longer feels current or conversion-ready for today's GCC market."
        )
    body = (
        f"{g}\n\n"
        f"{open_line}\n\n"
        f"{problem}\n\n"
        f"In this category, {angle}.\n\n"
        f"At DMC Creatives Studio we help Gulf businesses with {offer.lower()}. "
        "Happy to share a free one-page concept for your brand this week — no obligation.\n\n"
        f"{SIGN}"
    )
    return body


def site_hook(wrong: str, max_len: int = 120) -> str:
    w = (wrong or "").strip()
    if not w:
        return "your current site does not feel conversion-ready on mobile"
    if len(w) <= max_len:
        return w[0].lower() + w[1:] if w else w
    cut = w[: max_len - 1].rsplit(" ", 1)[0]
    return cut[0].lower() + cut[1:] + "…"


def whatsapp_msg(r: dict) -> str:
    company = r["company"]
    g = greeting(r.get("owner_or_dm") or "")
    offer = r["offer"]
    area = (r.get("locality") or "UAE").split(",")[0]
    industry = r.get("industry") or "Business"
    lead_type = (r.get("lead_type") or "").lower()
    hook = site_hook(r.get("wrong") or "")

    if lead_type == "new_business":
        opener = (
            f"I saw {company} launching in {area} — strong brand, but most enquiries still "
            f"land on WhatsApp/Instagram instead of a owned website."
        )
        pain = (
            "A simple 5-page site (About, Services, Gallery, Contact + WhatsApp button) "
            "helps you look established when buyers Google you."
        )
    elif industry == "Healthcare":
        opener = (
            f"I was comparing clinics in {area} and noticed {company}'s website — {hook}."
        )
        pain = (
            "Most patients shortlist 2–3 clinics on mobile before they call. "
            "A clean site with WhatsApp booking usually wins that first enquiry."
        )
    elif industry in {"Manufacturing", "Trading"}:
        opener = (
            f"I came across {company} while reviewing industrial suppliers in {area} — {hook}."
        )
        pain = (
            "RFQ buyers in the UAE usually pick suppliers with a clear product catalogue "
            "and a WhatsApp quote button — not a thin brochure page."
        )
    elif industry == "Automotive":
        opener = (
            f"I checked {company}'s site — {hook}."
        )
        pain = (
            "Most garage bookings in Dubai/RAK now start on Google or WhatsApp. "
            "A mobile site with service list + click-to-call helps you capture walk-ins."
        )
    elif industry == "Education":
        opener = (
            f"Parents comparing schools in {area} check the website on mobile before admissions calls — "
            f"{company}'s online presence could convert more enquiries."
        )
        pain = "Happy to show how a cleaner admissions page + WhatsApp enquiry could look."
    elif industry == "Services":
        opener = (
            f"I reviewed {company} in {area} — {hook}."
        )
        pain = (
            "Local service buyers trust a proper website over Gmail/listing pages "
            "when they compare vendors."
        )
    else:
        opener = (
            f"I was reviewing businesses in {area} and noticed {company}'s "
            "website looks outdated for today's Gulf market."
        )
        pain = "Buyers judge readiness from your website before they ever call."

    return (
        f"{g}\n\n"
        f"I'm Vaibhav from DMC Creatives Studio (India). {opener}\n\n"
        f"{pain}\n\n"
        f"We help with {offer.lower()}. Packages from ~AED 2,500 for SMEs.\n"
        "I can share a free one-page mockup this week — no obligation.\n\n"
        "Vaibhav Gurav\n"
        "DMC Creatives Studio\n"
        "hello@dmcstudio.in\n"
        "www.dmcstudio.in\n"
        "+91 83693 61785"
    )


def normalize_row(raw: dict) -> dict | None:
    company = (raw.get("company") or "").strip()
    if not company:
        return None
    email_raw = raw.get("email") or ""
    phone_raw = raw.get("phone") or ""
    emails = extract_emails(email_raw)
    mobiles = extract_mobiles(phone_raw, raw.get("locality") or "")
    if not emails and not has_any_phone(phone_raw) and not mobiles:
        return None
    priority = raw.get("priority") or "Medium"
    if priority not in {"High", "Medium", "Low"}:
        if "P1" in str(priority) or "P2" in str(priority) or "Hot" in str(priority):
            priority = "High"
        else:
            priority = "Medium"
    industry = (raw.get("industry") or "Business").strip()
    return {
        "company": company,
        "owner_or_dm": (raw.get("owner_or_dm") or raw.get("owner") or "").strip(),
        "designation": (raw.get("designation") or "").strip(),
        "email": emails[0] if emails else "",
        "emails": emails,
        "phone": phone_raw.strip(),
        "phones": mobiles,
        "website": (raw.get("website") or "").strip(),
        "wrong": (raw.get("wrong") or "").strip(),
        "industry": industry,
        "locality": (raw.get("locality") or "").strip(),
        "priority": priority,
        "offer": (raw.get("offer") or "Website redesign + lead capture").strip(),
        "value_note": (raw.get("value_note") or "").strip(),
        "source": (raw.get("source") or "").strip(),
        "lead_type": (raw.get("lead_type") or "outdated_site").strip(),
        "country": (raw.get("country") or "").strip(),
        "contact_week": int(raw.get("contact_week") or 0),
        "contact_rank": int(raw.get("contact_rank") or 0),
    }


def load_all() -> list[dict]:
    rows: list[dict] = []
    seen = set()
    if not GULF.exists():
        raise SystemExit(f"Missing {GULF}")
    data = json.loads(GULF.read_text(encoding="utf-8"))
    for raw in data:
        n = normalize_row(raw)
        if not n:
            continue
        key = n["company"].lower().strip()
        if key in seen:
            existing = next(x for x in rows if x["company"].lower().strip() == key)
            if not existing["email"] and n["email"]:
                rows.remove(existing)
                rows.append(n)
            continue
        seen.add(key)
        rows.append(n)
    rows.sort(key=lambda r: (
        0 if r["priority"] == "High" else 1,
        r["contact_rank"] if r.get("contact_rank") else 999,
        r["company"].lower(),
    ))
    return rows


def enrich(rows: list[dict]) -> list[dict]:
    for r in rows:
        r["subject"] = subject_line(r["company"], r["industry"], r["wrong"])
        r["email_body"] = email_body(r)
        r["whatsapp_msg"] = whatsapp_msg(r)
        r["wa_number"] = r["phones"][0] if r["phones"] else ""
        r["primary_phone_display"] = (
            display_phone(r["wa_number"]) if r["wa_number"] else (r["phone"] if has_any_phone(r["phone"]) else "")
        )
        if r["email"]:
            r["mailto"] = (
                "mailto:"
                + urllib.parse.quote(r["email"], safe="@.+_-")
                + "?"
                + urllib.parse.urlencode(
                    {"subject": r["subject"], "body": r["email_body"]},
                    quote_via=urllib.parse.quote,
                )
            )
        else:
            r["mailto"] = ""
        if r["wa_number"]:
            r["wa_link"] = (
                "https://wa.me/"
                + r["wa_number"]
                + "?text="
                + urllib.parse.quote(r["whatsapp_msg"])
            )
        else:
            r["wa_link"] = ""
    return rows


def write_excel(rows: list[dict]) -> Path:
    side = Side(style="thin", color="D0D0D0")
    border = Border(left=side, right=side, top=side, bottom=side)
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(bold=True, color="FFFFFF")
    prio_fill = {
        "High": PatternFill("solid", fgColor="C6EFCE"),
        "Medium": PatternFill("solid", fgColor="FFEDD5"),
        "Low": PatternFill("solid", fgColor="E0E7FF"),
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "Contactable Clients"
    headers = [
        "Priority",
        "Company",
        "Industry",
        "Decision Maker",
        "Designation",
        "Email",
        "Phone / WhatsApp",
        "Website",
        "Locality",
        "What's Wrong",
        "Offer",
        "Value Note",
        "Email Subject",
        "Personalized Email",
        "Personalized WhatsApp",
        "SEND EMAIL",
        "SEND WHATSAPP",
        "Source",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for i, r in enumerate(rows, 2):
        vals = [
            r["priority"],
            r["company"],
            r["industry"],
            r["owner_or_dm"],
            r["designation"],
            r["email"],
            r["primary_phone_display"] or r["phone"],
            r["website"],
            r["locality"],
            r["wrong"],
            r["offer"],
            r["value_note"],
            r["subject"],
            r["email_body"],
            r["whatsapp_msg"],
            "CLICK TO EMAIL" if r["mailto"] else "",
            "CLICK WHATSAPP" if r["wa_link"] else "",
            r["source"],
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i, c, v or "")
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c == 1 and v in prio_fill:
                cell.fill = prio_fill[v]
            if c == 6 and v and "@" in str(v):
                cell.hyperlink = f"mailto:{v}"
                cell.font = Font(color="0563C1", underline="single")
            if c == 16 and r["mailto"]:
                cell.hyperlink = r["mailto"]
                cell.font = Font(color="0563C1", underline="single")
            if c == 17 and r["wa_link"]:
                cell.hyperlink = r["wa_link"]
                cell.font = Font(color="128C7E", underline="single")
        ws.row_dimensions[i].height = 72

    widths = [10, 26, 16, 20, 18, 28, 18, 28, 20, 36, 26, 22, 34, 42, 36, 14, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:R{len(rows)+1}"

    ws2 = wb.create_sheet("High Priority")
    h2 = ["Company", "To", "Phone", "Subject", "Email Body", "WhatsApp", "CLICK EMAIL", "CLICK WHATSAPP"]
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(1, c, h)
        cell.fill = header_fill
        cell.font = header_font
    ri = 2
    for r in rows:
        if r["priority"] != "High":
            continue
        vals = [
            r["company"],
            r["email"] or "(use WhatsApp)",
            r["primary_phone_display"] or r["phone"],
            r["subject"],
            r["email_body"],
            r["whatsapp_msg"] if r["wa_link"] else "(no WhatsApp mobile)",
            "CLICK TO EMAIL" if r["mailto"] else "No email",
            "CLICK WHATSAPP" if r["wa_link"] else "No WA",
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(ri, c, v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
        if r["mailto"]:
            ws2.cell(ri, 7).hyperlink = r["mailto"]
            ws2.cell(ri, 7).font = Font(color="0563C1", underline="single")
        if r["wa_link"]:
            ws2.cell(ri, 8).hyperlink = r["wa_link"]
            ws2.cell(ri, 8).font = Font(color="128C7E", underline="single")
        ws2.row_dimensions[ri].height = 100
        ri += 1
    for i, w in enumerate([22, 28, 16, 34, 48, 36, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(XLSX_OUT)
    return XLSX_OUT


def write_html(rows: list[dict]) -> Path:
    contacts = [
        {
            "priority": r["priority"],
            "company": r["company"],
            "industry": r["industry"],
            "dm": r["owner_or_dm"],
            "des": r["designation"],
            "email": r["email"],
            "phone": r["primary_phone_display"] or r["phone"],
            "website": r["website"],
            "locality": r["locality"],
            "country": r.get("country") or "",
            "leadType": r.get("lead_type") or "",
            "contactWeek": r.get("contact_week") or 0,
            "contactRank": r.get("contact_rank") or 0,
            "wrong": r["wrong"],
            "offer": r["offer"],
            "valueNote": r["value_note"],
            "subject": r["subject"],
            "emailBody": r["email_body"],
            "whatsappMsg": r["whatsapp_msg"],
            "mailto": r["mailto"],
            "waLink": r["wa_link"],
            "waNumber": r["wa_number"],
        }
        for r in rows
    ]
    contacts_js = json.dumps(contacts, ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>DMC Client Console</title>
<link rel="preconnect" href="https://fonts.googleapis.com" />
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin />
<link href="https://fonts.googleapis.com/css2?family=Fira+Code:wght@500;600;700&family=Fira+Sans:wght@400;500;600;700&display=swap" rel="stylesheet" />
<style>
:root {{
  --color-primary: #2563EB;
  --color-on-primary: #FFFFFF;
  --color-secondary: #3B82F6;
  --color-accent: #059669;
  --color-background: #F8FAFC;
  --color-foreground: #0F172A;
  --color-muted: #F1F5FD;
  --color-border: #E4ECFC;
  --color-destructive: #DC2626;
  --color-ring: #2563EB;
  --text-muted: #64748B;
  --surface: #FFFFFF;
  --wa: #128C7E;
  --high: #059669;
  --med: #D97706;
  --space-2: 8px;
  --space-3: 12px;
  --space-4: 16px;
  --space-5: 24px;
  --radius: 10px;
  --font: "Fira Sans", system-ui, sans-serif;
  --mono: "Fira Code", ui-monospace, monospace;
  --shadow: 0 8px 24px rgba(15, 23, 42, 0.06);
}}
* {{ box-sizing: border-box; }}
html, body {{ height: 100%; }}
body {{
  margin: 0;
  font-family: var(--font);
  background: var(--color-background);
  color: var(--color-foreground);
  line-height: 1.45;
}}
button, a, input, select, textarea {{ font: inherit; }}
button, .btn, a.btn {{ cursor: pointer; transition: background-color 180ms ease, border-color 180ms ease, color 180ms ease; }}
button:focus-visible, a:focus-visible, input:focus-visible, select:focus-visible, textarea:focus-visible {{
  outline: 2px solid var(--color-ring);
  outline-offset: 2px;
}}
@media (prefers-reduced-motion: reduce) {{
  * {{ transition: none !important; animation: none !important; }}
}}
header {{
  padding: var(--space-4) var(--space-5) var(--space-3);
  background: var(--surface);
  border-bottom: 1px solid var(--color-border);
}}
header h1 {{
  margin: 0;
  font-size: 1.25rem;
  font-weight: 700;
  letter-spacing: -0.02em;
}}
header p {{
  margin: 4px 0 0;
  color: var(--text-muted);
  font-size: 0.88rem;
  max-width: 820px;
}}
.toolbar {{
  display: flex;
  flex-wrap: wrap;
  gap: var(--space-2);
  align-items: center;
  padding: var(--space-3) var(--space-5);
  background: rgba(248,250,252,0.96);
  border-bottom: 1px solid var(--color-border);
  position: sticky;
  top: 0;
  z-index: 20;
  backdrop-filter: blur(8px);
}}
input, select {{
  background: var(--surface);
  color: var(--color-foreground);
  border: 1px solid var(--color-border);
  border-radius: 8px;
  padding: 9px 12px;
  font-size: 0.88rem;
  min-height: 40px;
}}
input:hover, select:hover {{ border-color: #c7d7f8; }}
input[type=search] {{ min-width: 180px; flex: 1 1 220px; }}
select {{ flex: 0 1 auto; max-width: 100%; }}
.stats {{ display: flex; gap: 8px; flex-wrap: wrap; margin-left: auto; }}
.chip {{
  font-size: 0.75rem;
  font-weight: 600;
  padding: 6px 10px;
  border-radius: 999px;
  background: var(--color-muted);
  color: var(--color-foreground);
  border: 1px solid var(--color-border);
  font-family: var(--mono);
  white-space: nowrap;
}}
.layout {{
  display: grid;
  grid-template-columns: minmax(280px, 340px) 1fr;
  min-height: calc(100dvh - 150px);
}}
.list {{
  border-right: 1px solid var(--color-border);
  overflow: auto;
  max-height: calc(100dvh - 150px);
  background: var(--surface);
  -webkit-overflow-scrolling: touch;
}}
.item {{
  padding: 12px 16px;
  border-bottom: 1px solid var(--color-border);
  cursor: pointer;
  touch-action: manipulation;
}}
.item:hover {{ background: #F8FAFF; }}
.item.active {{
  background: #EFF6FF;
  box-shadow: inset 3px 0 0 var(--color-primary);
}}
.item .co {{ font-weight: 600; font-size: 0.92rem; word-break: break-word; }}
.item .meta {{ color: var(--text-muted); font-size: 0.78rem; margin-top: 3px; word-break: break-word; }}
.prio {{
  display: inline-block;
  font-size: 0.68rem;
  font-weight: 700;
  letter-spacing: .04em;
  padding: 2px 7px;
  border-radius: 4px;
  margin-right: 6px;
  text-transform: uppercase;
  font-family: var(--mono);
}}
.prio.High {{ background: #D1FAE5; color: #065F46; }}
.prio.Medium {{ background: #FFEDD5; color: #9A3412; }}
.prio.Low {{ background: #E0E7FF; color: #3730A3; }}
.detail {{
  padding: 20px 24px 48px;
  overflow: auto;
  max-height: calc(100dvh - 150px);
  -webkit-overflow-scrolling: touch;
}}
.back-btn {{
  display: none;
  appearance: none;
  border: 1px solid var(--color-border);
  background: var(--surface);
  color: var(--color-foreground);
  border-radius: 8px;
  min-height: 40px;
  padding: 0 12px;
  font-size: 0.85rem;
  font-weight: 600;
  margin-bottom: 12px;
  align-items: center;
  gap: 6px;
}}
.empty {{
  color: var(--text-muted);
  padding: 48px 24px;
  text-align: center;
}}
.detail h2 {{ margin: 0 0 4px; font-size: clamp(1.1rem, 2.5vw, 1.35rem); letter-spacing: -0.02em; word-break: break-word; }}
.sub {{ color: var(--text-muted); font-size: 0.88rem; margin-bottom: 14px; }}
.kv {{
  display: grid;
  grid-template-columns: 120px 1fr;
  gap: 6px 10px;
  font-size: 0.88rem;
  margin-bottom: 14px;
  padding: 12px 14px;
  background: var(--surface);
  border: 1px solid var(--color-border);
  border-radius: var(--radius);
}}
.kv span {{ min-width: 0; overflow-wrap: anywhere; }}
.kv span:nth-child(odd) {{ color: var(--text-muted); font-weight: 600; font-size: 0.78rem; text-transform: uppercase; letter-spacing: .03em; }}
.kv a {{ color: var(--color-primary); word-break: break-all; }}
.actions {{ display: flex; flex-wrap: wrap; gap: 8px; margin: 12px 0 10px; }}
.btn {{
  appearance: none;
  border: 1px solid transparent;
  border-radius: 8px;
  min-height: 44px;
  padding: 0 14px;
  font-size: 0.88rem;
  font-weight: 600;
  text-decoration: none;
  display: inline-flex;
  align-items: center;
  justify-content: center;
  gap: 8px;
  color: #fff;
  background: var(--color-muted);
  touch-action: manipulation;
}}
.btn svg {{ width: 16px; height: 16px; flex-shrink: 0; }}
.btn-email {{ background: var(--color-accent); }}
.btn-email:hover {{ background: #047857; }}
.btn-wa {{ background: var(--wa); }}
.btn-wa:hover {{ background: #0f766e; }}
.btn-copy {{ background: var(--color-primary); }}
.btn-copy:hover {{ background: #1d4ed8; }}
.btn-ghost {{ background: #fff; color: var(--color-foreground); border-color: var(--color-border); }}
.btn-ghost:hover {{ background: var(--color-muted); }}
.btn:disabled {{ opacity: .45; cursor: not-allowed; }}
.hint {{ font-size: 0.78rem; color: var(--text-muted); margin: 0 0 14px; line-height: 1.4; }}
.grid2 {{
  display: grid;
  grid-template-columns: 1fr 1fr;
  gap: 14px;
}}
.card {{
  background: var(--surface);
  border: 1px solid var(--color-border);
  border-radius: var(--radius);
  padding: 14px;
  box-shadow: var(--shadow);
  min-width: 0;
}}
.card h3 {{
  margin: 0 0 10px;
  font-size: 0.75rem;
  text-transform: uppercase;
  letter-spacing: .06em;
  color: var(--text-muted);
}}
.field {{ margin-bottom: 10px; }}
.field label {{
  display: block;
  font-size: 0.72rem;
  font-weight: 600;
  color: var(--text-muted);
  margin-bottom: 4px;
  text-transform: uppercase;
  letter-spacing: .04em;
}}
.field input, .field textarea {{
  width: 100%;
  max-width: 100%;
  background: #F8FAFF;
  border: 1px solid var(--color-border);
  color: var(--color-foreground);
  border-radius: 8px;
  padding: 10px;
  font-size: 16px;
  resize: vertical;
}}
textarea {{ min-height: 220px; line-height: 1.45; font-family: var(--font); }}
.toast {{
  position: fixed;
  bottom: 20px;
  right: 20px;
  left: auto;
  background: var(--color-accent);
  color: #fff;
  padding: 10px 16px;
  border-radius: 8px;
  font-size: 0.85rem;
  font-weight: 600;
  opacity: 0;
  pointer-events: none;
  transition: opacity .2s;
  z-index: 50;
  max-width: calc(100vw - 32px);
}}
.toast.show {{ opacity: 1; }}
.item {{
  opacity: 0;
  transform: translateY(8px);
  animation: fadeUp 280ms ease forwards;
}}
@keyframes fadeUp {{
  to {{ opacity: 1; transform: translateY(0); }}
}}

@media (max-width: 1100px) {{
  .grid2 {{ grid-template-columns: 1fr; }}
}}

@media (max-width: 900px) {{
  header {{ padding: 14px 16px 10px; }}
  header h1 {{ font-size: 1.1rem; }}
  header p {{ font-size: 0.8rem; }}
  .toolbar {{
    padding: 10px 12px;
    gap: 8px;
  }}
  input[type=search] {{
    flex: 1 1 100%;
    min-width: 0;
    width: 100%;
  }}
  select {{
    flex: 1 1 calc(50% - 4px);
    min-width: 0;
  }}
  .stats {{
    margin-left: 0;
    width: 100%;
  }}
  .layout {{
    grid-template-columns: 1fr;
    min-height: 0;
  }}
  .list, .detail {{
    max-height: none;
    border-right: none;
  }}
  .list {{
    max-height: min(42vh, 360px);
    border-bottom: 1px solid var(--color-border);
  }}
  .detail {{
    padding: 16px 14px 96px;
  }}
  body.mobile-detail .list {{ display: none; }}
  body.mobile-detail .detail {{
    min-height: calc(100dvh - 120px);
    padding-bottom: 120px;
  }}
  body.mobile-detail .back-btn {{ display: inline-flex; }}
  body:not(.mobile-detail) .detail .empty {{
    padding: 28px 16px;
  }}
  .kv {{
    grid-template-columns: 1fr;
    gap: 2px 0;
  }}
  .kv span:nth-child(odd) {{ margin-top: 8px; }}
  .kv span:nth-child(1) {{ margin-top: 0; }}
  .actions {{
    position: sticky;
    bottom: 0;
    z-index: 15;
    margin: 12px -14px 10px;
    padding: 10px 14px;
    background: rgba(248,250,252,0.96);
    border-top: 1px solid var(--color-border);
    backdrop-filter: blur(8px);
  }}
  .actions .btn {{
    flex: 1 1 calc(50% - 4px);
    min-width: 140px;
  }}
  textarea {{ min-height: 180px; }}
  .toast {{
    left: 50%;
    right: auto;
    bottom: 16px;
    transform: translateX(-50%);
  }}
}}

@media (max-width: 480px) {{
  select {{ flex: 1 1 100%; }}
  .chip {{ font-size: 0.68rem; padding: 5px 8px; }}
  .actions .btn {{ flex: 1 1 100%; }}
  .item {{ padding: 14px 14px; }}
}}
</style>
</head>
<body>
<header>
  <h1>DMC Client Console</h1>
  <p>UAE · Saudi Arabia · Qatar · Kuwait · Bahrain · Oman. New businesses that need a website, plus outdated sites ready for a rebuild. Signed as Vaibhav Gurav · DMC Creatives · hello@dmcstudio.in · +91 83693 61785.</p>
</header>
<div class="toolbar">
  <input type="search" id="q" placeholder="Search company, person, industry, city, country…" />
  <select id="priority">
    <option value="All">All priorities</option>
    <option value="High">High</option>
    <option value="Medium">Medium</option>
  </select>
  <select id="week">
    <option value="All">All leads</option>
    <option value="1" selected>UAE Week 1 (top 15)</option>
  </select>
  <select id="country"><option value="All">All countries</option></select>
  <select id="leadType">
    <option value="All">All lead types</option>
    <option value="new_business">New business</option>
    <option value="outdated_site">Outdated website</option>
  </select>
  <select id="industry"><option value="All">All industries</option></select>
  <select id="contact">
    <option value="All">Any contact</option>
    <option value="email">Has email</option>
    <option value="wa">Has WhatsApp</option>
    <option value="both">Email + WhatsApp</option>
  </select>
  <div class="stats">
    <span class="chip" id="countChip">0 shown</span>
    <span class="chip" id="highChip">0 High</span>
    <span class="chip" id="emailChip">0 email</span>
    <span class="chip" id="waChip">0 WhatsApp</span>
  </div>
</div>
<div class="layout">
  <div class="list" id="list"></div>
  <div class="detail" id="detail"><div class="empty">Select a company to view and send their personalized message.</div></div>
</div>
<div class="toast" id="toast">Copied</div>
<script>
const CONTACTS = {contacts_js};

const listEl = document.getElementById('list');
const detailEl = document.getElementById('detail');
const qEl = document.getElementById('q');
const pEl = document.getElementById('priority');
const weekEl = document.getElementById('week');
const countryEl = document.getElementById('country');
const typeEl = document.getElementById('leadType');
const iEl = document.getElementById('industry');
const cEl = document.getElementById('contact');
const toast = document.getElementById('toast');
let selected = null;

[...new Set(CONTACTS.map(c => c.industry).filter(Boolean))].sort().forEach(c => {{
  const o = document.createElement('option');
  o.value = c; o.textContent = c; iEl.appendChild(o);
}});
[...new Set(CONTACTS.map(c => c.country || (c.locality || '').split(',').pop().trim()).filter(Boolean))].sort().forEach(c => {{
  const o = document.createElement('option');
  o.value = c; o.textContent = c; countryEl.appendChild(o);
}});

function toastMsg(msg) {{
  toast.textContent = msg;
  toast.classList.add('show');
  setTimeout(() => toast.classList.remove('show'), 1800);
}}

function filtered() {{
  const q = qEl.value.trim().toLowerCase();
  return CONTACTS.filter(c => {{
    if (pEl.value !== 'All' && c.priority !== pEl.value) return false;
    if (iEl.value !== 'All' && c.industry !== iEl.value) return false;
    if (weekEl.value !== 'All' && String(c.contactWeek) !== weekEl.value) return false;
    if (countryEl.value !== 'All' && c.country !== countryEl.value && !(c.locality || '').includes(countryEl.value)) return false;
    if (typeEl.value !== 'All' && c.leadType !== typeEl.value) return false;
    if (cEl.value === 'email' && !c.email) return false;
    if (cEl.value === 'wa' && !c.waNumber) return false;
    if (cEl.value === 'both' && !(c.email && c.waNumber)) return false;
    if (!q) return true;
    return `${{c.company}} ${{c.dm}} ${{c.des}} ${{c.industry}} ${{c.email}} ${{c.locality}} ${{c.country}} ${{c.offer}} ${{c.leadType}}`.toLowerCase().includes(q);
  }});
}}

function updateStats(rows) {{
  document.getElementById('countChip').textContent = rows.length + ' shown';
  document.getElementById('highChip').textContent = rows.filter(r => r.priority === 'High').length + ' High';
  document.getElementById('emailChip').textContent = rows.filter(r => r.email).length + ' email';
  document.getElementById('waChip').textContent = rows.filter(r => r.waNumber).length + ' WhatsApp';
}}

function isMobile() {{
  return window.matchMedia('(max-width: 900px)').matches;
}}

function setMobilePane(openDetail) {{
  document.body.classList.toggle('mobile-detail', Boolean(openDetail) && isMobile());
}}

function renderList() {{
  const rows = filtered();
  updateStats(rows);
  listEl.innerHTML = rows.map((c, idx) => `
    <div class="item ${{selected && selected.company === c.company ? 'active' : ''}}" data-company="${{c.company.replace(/"/g,'&quot;')}}" style="animation-delay:${{Math.min(idx, 12) * 30}}ms">
      <div class="co">${{c.contactRank ? `<span class="prio High">#${{c.contactRank}}</span>` : ''}}<span class="prio ${{c.priority}}">${{c.priority}}</span>${{esc(c.company)}}</div>
      <div class="meta">${{esc(c.industry)}} · ${{esc(c.leadType === 'new_business' ? 'New business' : 'Outdated site')}} · ${{esc(c.dm) || 'No named contact'}} · ${{esc(c.locality)}}</div>
    </div>
  `).join('');
  listEl.querySelectorAll('.item').forEach(el => {{
    el.addEventListener('click', () => {{
      selected = CONTACTS.find(x => x.company === el.dataset.company);
      renderList();
      renderDetail();
      setMobilePane(true);
      if (isMobile()) detailEl.scrollIntoView({{ behavior: 'smooth', block: 'start' }});
    }});
  }});
}}

function esc(s) {{
  return String(s || '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}}

function iconMail() {{
  return '<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M4 6h16v12H4z"/><path d="m4 7 8 6 8-6"/></svg>';
}}
function iconWa() {{
  return '<svg viewBox="0 0 24 24" fill="currentColor"><path d="M12 2a10 10 0 0 0-8.7 14.9L2 22l5.3-1.4A10 10 0 1 0 12 2zm0 18a8 8 0 0 1-4.1-1.1l-.3-.2-3 .8.8-2.9-.2-.3A8 8 0 1 1 12 20zm4.4-5.9c-.2-.1-1.4-.7-1.6-.8s-.4-.1-.5.1-.6.8-.8 1-.3.2-.5.1a6.5 6.5 0 0 1-3.2-2.8c-.2-.4.2-.4.6-1.3.1-.1 0-.3 0-.4s-.5-1.2-.7-1.6-.4-.4-.5-.4h-.4c-.2 0-.4.1-.6.3a2 2 0 0 0-.6 1.5 3.5 3.5 0 0 0 .7 1.8 8 8 0 0 0 3.1 2.9 10.5 10.5 0 0 0 2.3.9 2.2 2.2 0 0 0 1.6.1 2.7 2.7 0 0 0 1.1-1c.1-.3.1-.5.1-.5s-.1-.1-.3-.2z"/></svg>';
}}

function renderDetail() {{
  if (!selected) {{
    detailEl.innerHTML = '<div class="empty">Select a company to view and send their personalized message.</div>';
    setMobilePane(false);
    return;
  }}
  const c = selected;
  detailEl.innerHTML = `
    <button type="button" class="back-btn" id="backList" aria-label="Back to list">← Back to list</button>
    <h2>${{esc(c.company)}}</h2>
    <div class="sub"><span class="prio ${{c.priority}}">${{c.priority}}</span> ${{esc(c.industry)}} · ${{esc(c.locality)}}</div>
    <div class="kv">
      <span>Decision maker</span><span>${{esc(c.dm) || '—'}} ${{c.des ? '· ' + esc(c.des) : ''}}</span>
      <span>Email</span><span>${{esc(c.email) || 'No public email — use WhatsApp'}}</span>
      <span>Phone</span><span style="font-family:var(--mono)">${{esc(c.phone) || '—'}}</span>
      <span>Website</span><span>${{c.website && c.website.startsWith('http') ? `<a href="${{esc(c.website)}}" target="_blank" rel="noopener">${{esc(c.website)}}</a>` : esc(c.website) || '—'}}</span>
      <span>What's wrong</span><span>${{esc(c.wrong)}}</span>
      <span>Offer</span><span>${{esc(c.offer)}}</span>
    </div>
    <div class="actions">
      ${{c.mailto ? `<a class="btn btn-email" id="sendEmail" href="${{esc(c.mailto)}}">${{iconMail()}} Send Email (1-click)</a>` : `<button class="btn btn-email" disabled>No email on file</button>`}}
      ${{c.waLink ? `<a class="btn btn-wa" id="sendWa" href="${{esc(c.waLink)}}" target="_blank" rel="noopener">${{iconWa()}} WhatsApp (1-click)</a>` : `<button class="btn btn-wa" disabled>No WhatsApp number</button>`}}
      <button class="btn btn-copy" id="copyEmail">Copy Email</button>
      <button class="btn btn-copy" id="copyWa">Copy WhatsApp</button>
      <button class="btn btn-ghost" id="copyAll">Copy Subject + Email</button>
    </div>
    <p class="hint">Email opens Outlook/Gmail with draft filled. WhatsApp opens wa.me with the message ready (mobile numbers only). Edit drafts below — links rebuild live.</p>
    <div class="grid2">
      <div class="card">
        <h3>Email draft</h3>
        <div class="field"><label>Subject</label><input id="subj" value="${{esc(c.subject)}}" /></div>
        <div class="field"><label>Body (editable)</label><textarea id="body">${{esc(c.emailBody)}}</textarea></div>
      </div>
      <div class="card">
        <h3>WhatsApp message</h3>
        <div class="field"><label>Message (editable)</label><textarea id="wamsg">${{esc(c.whatsappMsg)}}</textarea></div>
        <p class="hint">${{c.waNumber ? 'Will send to +' + esc(c.waNumber) : 'No usable mobile — copy and send manually if you have another number.'}}</p>
      </div>
    </div>
  `;

  const back = document.getElementById('backList');
  if (back) {{
    back.onclick = () => {{
      setMobilePane(false);
      listEl.scrollIntoView({{ behavior: 'smooth', block: 'start' }});
    }};
  }}

  const subj = document.getElementById('subj');
  const body = document.getElementById('body');
  const wamsg = document.getElementById('wamsg');
  const send = document.getElementById('sendEmail');
  const sendWa = document.getElementById('sendWa');

  function rebuildMailto() {{
    if (!send || !c.email) return;
    send.href = 'mailto:' + encodeURIComponent(c.email).replace(/%40/g,'@')
      + '?subject=' + encodeURIComponent(subj.value)
      + '&body=' + encodeURIComponent(body.value);
  }}
  function rebuildWa() {{
    if (!sendWa || !c.waNumber) return;
    sendWa.href = 'https://wa.me/' + c.waNumber + '?text=' + encodeURIComponent(wamsg.value);
  }}
  subj.addEventListener('input', rebuildMailto);
  body.addEventListener('input', rebuildMailto);
  wamsg.addEventListener('input', rebuildWa);

  document.getElementById('copyEmail').onclick = () => {{
    navigator.clipboard.writeText(body.value); toastMsg('Email body copied');
  }};
  document.getElementById('copyWa').onclick = () => {{
    navigator.clipboard.writeText(wamsg.value); toastMsg('WhatsApp message copied');
  }};
  document.getElementById('copyAll').onclick = () => {{
    navigator.clipboard.writeText('Subject: ' + subj.value + '\\n\\n' + body.value);
    toastMsg('Subject + email copied');
  }};
}}

qEl.addEventListener('input', renderList);
pEl.addEventListener('change', renderList);
weekEl.addEventListener('change', renderList);
countryEl.addEventListener('change', renderList);
typeEl.addEventListener('change', renderList);
iEl.addEventListener('change', renderList);
cEl.addEventListener('change', renderList);
window.addEventListener('resize', () => {{
  if (!isMobile()) document.body.classList.remove('mobile-detail');
  else if (selected && document.body.classList.contains('mobile-detail')) setMobilePane(true);
}});
renderList();
if (!isMobile()) {{
  const firstWeek = CONTACTS.find(c => c.contactWeek === 1 && c.contactRank === 1)
    || CONTACTS.find(c => c.contactWeek === 1)
    || CONTACTS.find(c => c.priority === 'High')
    || CONTACTS[0];
  if (firstWeek) {{ selected = firstWeek; renderList(); renderDetail(); }}
}} else {{
  renderDetail();
}}
</script>
</body>
</html>
"""
    HTML_OUT.write_text(html, encoding="utf-8")
    return HTML_OUT


def main() -> None:
    rows = enrich(load_all())
    JSON_OUT.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    write_excel(rows)
    write_html(rows)
    with_email = sum(1 for r in rows if r["email"])
    with_wa = sum(1 for r in rows if r["wa_number"])
    high = sum(1 for r in rows if r["priority"] == "High")
    print(f"Contactable: {len(rows)}")
    print(f"High: {high} | Email: {with_email} | WhatsApp: {with_wa}")
    print(f"Wrote {JSON_OUT.name}")
    print(f"Wrote {XLSX_OUT.name}")
    print(f"Wrote {HTML_OUT.name}")
    assert len(rows) >= 40, f"Expected 40+ contactable Gulf leads, got {len(rows)}"


if __name__ == "__main__":
    main()
